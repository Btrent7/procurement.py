#New Part Number Script
import pandas as pd
import openpyxl as op
from pricepy import markup #Unique Function for list_price value
from datetime import date
import base64
from email.mime.text import MIMEText
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from requests import HTTPError 


# Excel File Path
newPart_form = "C:/Users//NewPartNumber_Form.xlsx"
newPart_table = "C:/Users//NewPartNumber_Table.xlsx"


#Load Form Workbook (openpyxl)
wb_form = op.load_workbook(newPart_form)
form = wb_form["newPart"]


#Form Variables
vnd_name    = form["B2"].value.upper()
vnd_id      = form["B3"].value.upper()
sku         = form["B4"].value.upper()
detail      = form["B5"].value.upper()
tpp         = float(form["B10"].value)
cat_code    = form["B11"].value
site        = form["B12"].value.upper()
request     = form["B13"].value

#Date Variable
today = date.today()

#Item Description
item_descr = (f"{vnd_name},#{sku},{detail}")

#List Price Function
list_price = markup(category_code= cat_code, tpp_value= tpp)
print(f"""Form accessed, markup applied: {today}""")


#Load Table Workbook (openpyxl)
wb_table = op.load_workbook(newPart_table)
table = wb_table["699_Table"]


#Select Next Blank Row in Table
last_row = 1
while table.cell(row=last_row, column=1).value is not None:
    last_row += 1


#Select Previous Row (for previous part number)
prev_pn_cell = table.cell(row= last_row - 1, column= 1).value
previous_pn = int(prev_pn_cell)


#Create New Part Number
new_pn = previous_pn + 1


#Fill Next Blank Row on Table
next_row = last_row
table.cell(row= next_row, column= 1, value = new_pn)
table.cell(row= next_row, column= 2, value= sku)
table.cell(row= next_row, column= 3, value= item_descr)
table.cell(row= next_row, column= 4, value= today)
table.cell(row= next_row, column= 5, value= site)
table.cell(row= next_row, column= 6, value= tpp)
table.cell(row= next_row, column= 7, value= cat_code)
table.cell(row= next_row, column= 11, value= list_price)

wb_table.save(newPart_table)
wb_table.close()


#Email Message Variable
email_body = f"""
New PN: {new_pn}
Desc: {item_descr}

Request: {request}

Vnd Name: {vnd_name}
Vnd ID: {vnd_id}
SKU: {sku} 
Detail: {detail} 
TPP: $ {tpp} 
Site: {site}

Thanks,
.py"""


#Gmail API Variables
SCOPES = [ "https://www.googleapis.com/auth/gmail.send"]
flow = InstalledAppFlow.from_client_secrets_file(r'C:/Users/btrent//creds.json', SCOPES)
creds = flow.run_local_server(port=0)
service = build('gmail', 'v1', credentials=creds)


#Email Detail Variables
message = MIMEText(email_body)
message['to'] = 'btrent@email.com'
message['subject'] = '@noreply'


#Send Email w/ New Part Number Information
create_message =  {'raw': base64.urlsafe_b64encode(message.as_bytes()).decode()}

try:
    result = (service.users().messages().send(userId="me", body=create_message).execute())
    print(f"""
          Message Delivered!""")
except HTTPError as error:
    print(f'An error occured {error}')
    result = None    



print(f"""
New PN: {new_pn}
VNDID:  {vnd_id}
Item:   {item_descr}
TPP:    {tpp}
List:   {list_price}
Site:   {site}
""")

#THE END
