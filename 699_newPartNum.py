from openpyxl import load_workbook

# Excel File Path
newPart_form = "C:/Users/Btrent/Documents/Python/NewPartNumber/NewPartNumber_Form.xlsx"
newPart_table = "C:/Users/Btrent/Documents/Python/NewPartNumber/NewPartNumber_Table.xlsx"

#Load Excel File (newPart_form)
wb_form = load_workbook(newPart_form)

#Select Sheet (699_)
ws_form = wb_form["699_NP#"]

#newPart_Form Variables
vnd_name = ws_form["B2"].value
sku = ws_form["B4"].value
detail = ws_form["B5"].value
tpp = ws_form["B10"]
cat_code = ws_form["B11"].value
site = ws_form["B12"].value
item_description = (f"{vnd_name}, #{sku}, {detail}")

#List Price Categories
if cat_code in ["501A", "501B", "501C"]:
    markup = ((1.2 * tpp.value) / 0.074)
elif cat_code in ["501D", "503G", "503H", "503I"]:
    markup = ((1.2 * tpp.value) / 0.085)
elif cat_code in ["501F", "501G", "501H", "501I"]:
    markup = ((1.2 * tpp.value) / 0.087)
elif cat_code in ["502S", "502T", "502U"]:
    markup = ((1.2 * tpp.value) / 0.090)
elif cat_code in ["503A", "503B", "503C", "503D", "503E", "503F"]:
    markup = ((1.2 * tpp.value) / 0.095)
elif cat_code in ["502G", "502H", "502I", "502J", "511C"]:
    markup = ((1.2 * tpp.value) / 0.100)
elif cat_code in ["515A", "515B"]:
    markup = ((1.2 * tpp.value) / 0.125)
elif cat_code == "514A":
    markup = ((1.2 * tpp.value) / 0.135)
elif cat_code in ["505A", "505B", "505C", "505D", "505E", "505F", "505G", "505H", "505I"]:
    markup = ((1.2 * tpp.value) / 0.155)
elif cat_code == "502P":
    markup = ((1.2 * tpp.value) / 0.160)
elif cat_code in ["502N", "502Q", "502R", "503K"]:
    markup = ((1.2 * tpp.value) / 0.165)
elif cat_code in ["502A", "502B", "502C", "502D", "502E"]:
    markup = ((1.2 * tpp.value) / 0.168)
elif cat_code == "501E":
    markup = ((1.2 * tpp.value) / 0.170)    
elif cat_code in ["511B", "511M"]:
    markup = ((1.2 * tpp.value) / 0.175)
elif cat_code in ["502K", "502L", "504A", "504B", "504C", "504D", "504E", "504F", "504G", "504H", "504I"]:
    markup = ((1.2 * tpp.value) / 0.180)
elif cat_code in ["503L", "503M", "503S", "511D", "511E"]:
    markup = ((1.2 * tpp.value) / 0.185)
elif cat_code == "511J":
    markup = ((1.2 * tpp.value) / 0.190)
elif cat_code in ["511A", "511Y"]:
    markup = ((1.2 * tpp.value) / 0.198)
elif cat_code == "511L":
    markup = ((1.2 * tpp.value) / 0.200)
elif cat_code in ["506A", "506B"]:
    markup = ((1.2 * tpp.value) / 0.220)
elif cat_code == "508A":
    markup = ((1.2 * tpp.value) / 0.230)
elif cat_code in ["442A", "511T"]:
    markup = ((1.2 * tpp.value) / 0.245)
elif cat_code in ["507A", "536A"]:
    markup = ((1.2 * tpp.value) / 0.260)
elif cat_code in ["510A", "510D"]:
    markup = ((1.2 * tpp.value) / 0.265)
elif cat_code in ["465A", "510Z"]:
    markup = ((1.2 * tpp.value) / 0.270)
elif cat_code in ["513A", "513B", "513C", "513D", "513E", "513F"]:
    markup = ((1.2 * tpp.value) / 0.275)
elif cat_code in ["450A", "511K", "518A"]:
    markup = ((1.2 * tpp.value) / 0.280)
elif cat_code == "507B":
    markup = ((1.2 * tpp.value) / 0.285)
elif cat_code in ["257A", "257B", "257C", "257F", "257G", "257H", "500B", "500C", "509A", "509B", "509C", "511H", "511I", "515C"]:
    markup = ((1.2 * tpp.value) / 0.300)
elif cat_code == "507C":
    markup = ((1.2 * tpp.value) / 0.310)
elif cat_code == "343B":
    markup = ((1.2 * tpp.value) / 0.320)
elif cat_code == "509D":
    markup = ((1.2 * tpp.value) / 0.340)
elif cat_code == "444A":
    markup = ((1.2 * tpp.value) / 0.350)
elif cat_code in ["519A", "519B", "519C", "519D"]:
    markup = ((1.2 * tpp.value) / 0.400)
elif cat_code == "506M":
    markup = ((1.2 * tpp.value) / 0.440)
elif cat_code in ["500A", "502M"]:
    markup = ((1.2 * tpp.value) / 0.500)
elif cat_code == "F":
    markup = ((1 * tpp.value) / 1.000)
else:
    print(f"Undefined Item Category: '{cat_code}'")
    wb_form.close()
    exit()

#List Price Formula & Variable
if isinstance(tpp.value, (int, float)):
    list_Price = markup

#Load Excel File (newPart_table)
wb_table = load_workbook(newPart_table)

#Select Sheet (699_)
ws_table = wb_table["699_Table"]

#Find Next Row
next_row = ws_table.max_row

#New Row Entry Data
next_row = 1
while ws_table.cell(row=next_row, column=2).value is not None:
    next_row += 1


ws_table.cell(row=next_row, column = 2, value = item_description)
ws_table.cell(row=next_row, column = 3, value=vnd_name)
ws_table.cell(row=next_row, column = 4, value = sku)
# ws_table.cell(row=next_row, column = 5, value = [DATE]
ws_table.cell(row=next_row, column = 6, value = cat_code)
ws_table.cell(row=next_row, column = 7, value = tpp.value)
ws_table.cell(row=next_row, column = 8, value = list_Price)
ws_table.cell(row=next_row, column = 9, value = site)

wb_table.save(newPart_table)
wb_table.close()

#Print Vendor Name & Item Description
print(f"Item Description: {item_description}")
print(f"TPP Price: {tpp.value}")
print(f"List Price: {list_Price}")
print(f"Site: {site}")
